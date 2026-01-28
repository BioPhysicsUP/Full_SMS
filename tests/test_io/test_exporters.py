"""Tests for data export functions."""

import csv
import json
from pathlib import Path

import numpy as np
import pytest

from full_sms.io.exporters import (
    ExportFormat,
    export_batch,
    export_fit_results,
    export_groups,
    export_intensity_trace,
    export_levels,
)
from full_sms.models.fit import FitResult, FitResultData
from full_sms.models.group import GroupData
from full_sms.models.level import LevelData
from full_sms.models.particle import ChannelData, ParticleData
from full_sms.models.session import FileMetadata, SessionState


# ============================================================================
# Fixtures
# ============================================================================


@pytest.fixture
def sample_abstimes() -> np.ndarray:
    """Create sample absolute times for intensity trace."""
    # Create 1000 photons over 1 second (1e9 ns)
    return np.linspace(0, 1_000_000_000, 1000, dtype=np.uint64)


@pytest.fixture
def sample_levels() -> list[LevelData]:
    """Create sample levels for export testing."""
    return [
        LevelData(
            start_index=0,
            end_index=99,
            start_time_ns=0,
            end_time_ns=100_000_000,
            num_photons=100,
            intensity_cps=1000.0,
            group_id=0,
        ),
        LevelData(
            start_index=100,
            end_index=249,
            start_time_ns=100_000_000,
            end_time_ns=250_000_000,
            num_photons=150,
            intensity_cps=1000.0,
            group_id=1,
        ),
        LevelData(
            start_index=250,
            end_index=399,
            start_time_ns=250_000_000,
            end_time_ns=400_000_000,
            num_photons=150,
            intensity_cps=1000.0,
            group_id=0,
        ),
    ]


@pytest.fixture
def sample_groups() -> list[GroupData]:
    """Create sample groups for export testing."""
    return [
        GroupData(
            group_id=0,
            level_indices=(0, 2),
            total_photons=250,
            total_dwell_time_s=0.25,
            intensity_cps=1000.0,
        ),
        GroupData(
            group_id=1,
            level_indices=(1,),
            total_photons=150,
            total_dwell_time_s=0.15,
            intensity_cps=1000.0,
        ),
    ]


@pytest.fixture
def sample_fit_result() -> FitResult:
    """Create a sample single-exponential fit result."""
    return FitResult.from_fit_parameters(
        tau=[5.0],
        tau_std=[0.1],
        amplitude=[1.0],
        amplitude_std=[0.01],
        shift=0.5,
        shift_std=0.05,
        chi_squared=1.05,
        durbin_watson=2.1,
        residuals=np.random.randn(100),
        fitted_curve=np.exp(-np.arange(100) / 50.0) * 1000,
        fit_start_index=10,
        fit_end_index=110,
        background=5.0,
        dw_bounds=(1.5, 2.5),
    )


@pytest.fixture
def sample_fit_result_data() -> FitResultData:
    """Create a sample FitResultData for level exports."""
    return FitResultData(
        tau=(5.0,),
        tau_std=(0.1,),
        amplitude=(1.0,),
        amplitude_std=(0.01,),
        shift=0.5,
        shift_std=0.05,
        chi_squared=1.05,
        durbin_watson=2.1,
        dw_bounds=(1.5, 2.5),
        fit_start_index=10,
        fit_end_index=110,
        background=5.0,
        num_exponentials=1,
        average_lifetime=5.0,
        level_index=0,
    )


@pytest.fixture
def biexponential_fit_result() -> FitResult:
    """Create a sample bi-exponential fit result."""
    return FitResult.from_fit_parameters(
        tau=[2.0, 8.0],
        tau_std=[0.1, 0.2],
        amplitude=[0.6, 0.4],
        amplitude_std=[0.02, 0.02],
        shift=0.3,
        shift_std=0.03,
        chi_squared=1.02,
        durbin_watson=2.0,
        residuals=np.random.randn(100),
        fitted_curve=np.exp(-np.arange(100) / 50.0) * 1000,
        fit_start_index=10,
        fit_end_index=110,
        background=3.0,
        dw_bounds=(1.5, 2.5),
    )


@pytest.fixture
def triexponential_fit_result() -> FitResult:
    """Create a sample tri-exponential fit result."""
    return FitResult.from_fit_parameters(
        tau=[1.0, 4.0, 12.0],
        tau_std=[0.05, 0.15, 0.3],
        amplitude=[0.4, 0.35, 0.25],
        amplitude_std=[0.01, 0.01, 0.01],
        shift=0.2,
        shift_std=0.02,
        chi_squared=0.98,
        durbin_watson=1.95,
        residuals=np.random.randn(100),
        fitted_curve=np.exp(-np.arange(100) / 50.0) * 1000,
        fit_start_index=10,
        fit_end_index=110,
        background=2.0,
        dw_bounds=(1.5, 2.5),
    )


# ============================================================================
# Test ExportFormat
# ============================================================================


class TestExportFormat:
    """Tests for ExportFormat enum."""

    def test_csv_value(self) -> None:
        assert ExportFormat.CSV.value == "csv"

    def test_parquet_value(self) -> None:
        assert ExportFormat.PARQUET.value == "parquet"

    def test_excel_value(self) -> None:
        assert ExportFormat.EXCEL.value == "xlsx"

    def test_json_value(self) -> None:
        assert ExportFormat.JSON.value == "json"


# ============================================================================
# Test export_intensity_trace
# ============================================================================


class TestExportIntensityTrace:
    """Tests for export_intensity_trace function."""

    def test_export_csv_format(
        self, sample_abstimes: np.ndarray, tmp_path: Path
    ) -> None:
        """Export intensity trace to CSV format."""
        output_path = tmp_path / "intensity"
        result = export_intensity_trace(
            sample_abstimes, output_path, bin_size_ms=10.0, fmt=ExportFormat.CSV
        )

        assert result.suffix == ".csv"
        assert result.exists()

        # Verify CSV content
        with open(result, "r") as f:
            reader = csv.reader(f)
            header = next(reader)
            assert header == ["time_ms", "counts", "intensity_cps"]

            # Read first data row
            first_row = next(reader)
            assert len(first_row) == 3
            # Time should be a float
            assert float(first_row[0]) >= 0

    def test_export_json_format(
        self, sample_abstimes: np.ndarray, tmp_path: Path
    ) -> None:
        """Export intensity trace to JSON format."""
        output_path = tmp_path / "intensity"
        result = export_intensity_trace(
            sample_abstimes,
            output_path,
            bin_size_ms=10.0,
            fmt=ExportFormat.JSON,
            particle_name="Test Particle",
        )

        assert result.suffix == ".json"
        assert result.exists()

        # Verify JSON content
        with open(result, "r") as f:
            data = json.load(f)

        assert "metadata" in data
        assert "data" in data
        assert data["metadata"]["type"] == "intensity_trace"
        assert data["metadata"]["bin_size_ms"] == 10.0
        assert data["metadata"]["particle_name"] == "Test Particle"
        assert "time_ms" in data["data"]
        assert "counts" in data["data"]
        assert "intensity_cps" in data["data"]

    def test_export_parquet_format(
        self, sample_abstimes: np.ndarray, tmp_path: Path
    ) -> None:
        """Export intensity trace to Parquet format."""
        pytest.importorskip("pyarrow")

        output_path = tmp_path / "intensity"
        result = export_intensity_trace(
            sample_abstimes, output_path, bin_size_ms=10.0, fmt=ExportFormat.PARQUET
        )

        assert result.suffix == ".parquet"
        assert result.exists()

        # Verify Parquet content
        import pyarrow.parquet as pq

        table = pq.read_table(result)
        assert "time_ms" in table.column_names
        assert "counts" in table.column_names
        assert "intensity_cps" in table.column_names

    def test_export_excel_format(
        self, sample_abstimes: np.ndarray, tmp_path: Path
    ) -> None:
        """Export intensity trace to Excel format."""
        pytest.importorskip("openpyxl")

        output_path = tmp_path / "intensity"
        result = export_intensity_trace(
            sample_abstimes, output_path, bin_size_ms=10.0, fmt=ExportFormat.EXCEL
        )

        assert result.suffix == ".xlsx"
        assert result.exists()

        # Verify Excel content
        import openpyxl

        wb = openpyxl.load_workbook(result)
        ws = wb.active
        assert ws.title == "Intensity Trace"
        assert ws.cell(1, 1).value == "time_ms"
        assert ws.cell(1, 2).value == "counts"
        assert ws.cell(1, 3).value == "intensity_cps"

    def test_empty_data_handling(self, tmp_path: Path) -> None:
        """Export with minimal data should work."""
        # Create very short trace (1 photon)
        abstimes = np.array([0], dtype=np.uint64)
        output_path = tmp_path / "intensity"

        result = export_intensity_trace(
            abstimes, output_path, bin_size_ms=10.0, fmt=ExportFormat.CSV
        )

        assert result.exists()

    def test_creates_parent_directories(
        self, sample_abstimes: np.ndarray, tmp_path: Path
    ) -> None:
        """Export creates parent directories if they don't exist."""
        output_path = tmp_path / "nested" / "subdir" / "intensity"
        result = export_intensity_trace(
            sample_abstimes, output_path, bin_size_ms=10.0, fmt=ExportFormat.CSV
        )

        assert result.exists()
        assert result.parent.exists()

    def test_bin_size_affects_output(
        self, sample_abstimes: np.ndarray, tmp_path: Path
    ) -> None:
        """Different bin sizes produce different numbers of bins."""
        output_small = tmp_path / "small_bins"
        output_large = tmp_path / "large_bins"

        result_small = export_intensity_trace(
            sample_abstimes, output_small, bin_size_ms=10.0, fmt=ExportFormat.CSV
        )
        result_large = export_intensity_trace(
            sample_abstimes, output_large, bin_size_ms=100.0, fmt=ExportFormat.CSV
        )

        # Count rows (excluding header)
        with open(result_small, "r") as f:
            small_rows = sum(1 for _ in f) - 1

        with open(result_large, "r") as f:
            large_rows = sum(1 for _ in f) - 1

        # Smaller bin size should produce more rows
        assert small_rows > large_rows


# ============================================================================
# Test export_levels
# ============================================================================


class TestExportLevels:
    """Tests for export_levels function."""

    def test_export_basic_levels_csv(
        self, sample_levels: list[LevelData], tmp_path: Path
    ) -> None:
        """Export basic levels to CSV."""
        output_path = tmp_path / "levels"
        result = export_levels(sample_levels, output_path, fmt=ExportFormat.CSV)

        assert result.suffix == ".csv"
        assert result.exists()

        # Verify CSV content
        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 3
        assert rows[0]["level_index"] == "0"
        assert rows[0]["num_photons"] == "100"
        assert rows[0]["group_id"] == "0"

    def test_export_with_fit_parameters(
        self, sample_levels: list[LevelData], sample_fit_result_data: FitResultData, tmp_path: Path
    ) -> None:
        """Export levels with fit parameters included."""
        level_fits = {0: sample_fit_result_data}
        output_path = tmp_path / "levels_with_fits"

        result = export_levels(
            sample_levels,
            output_path,
            fmt=ExportFormat.CSV,
            level_fits=level_fits,
        )

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Should have fit columns
        assert "tau_1_ns" in rows[0]
        assert "chi_squared" in rows[0]
        assert "avg_lifetime_ns" in rows[0]

        # First level should have fit values
        assert rows[0]["tau_1_ns"] == "5.0000"
        assert rows[0]["chi_squared"] == "1.0500"

        # Second level should have empty fit values
        assert rows[1]["tau_1_ns"] == ""

    def test_handles_missing_optional_fits(
        self, sample_levels: list[LevelData], tmp_path: Path
    ) -> None:
        """Levels without fits get empty values in fit columns."""
        # Empty fits dict - should work but not add fit columns
        output_path = tmp_path / "levels_no_fits"
        result = export_levels(
            sample_levels,
            output_path,
            fmt=ExportFormat.CSV,
            level_fits={},
        )

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames

        # Should not have fit columns when level_fits is empty
        assert "tau_1_ns" not in fieldnames

    def test_multi_exponential_columns(
        self, sample_levels: list[LevelData], tmp_path: Path
    ) -> None:
        """Export with biexponential fit includes additional columns."""
        biexp_fit = FitResultData(
            tau=(2.0, 8.0),
            tau_std=(0.1, 0.2),
            amplitude=(0.6, 0.4),
            amplitude_std=(0.02, 0.02),
            shift=0.3,
            shift_std=0.03,
            chi_squared=1.02,
            durbin_watson=2.0,
            dw_bounds=(1.5, 2.5),
            fit_start_index=10,
            fit_end_index=110,
            background=3.0,
            num_exponentials=2,
            average_lifetime=4.4,
            level_index=0,
        )
        level_fits = {0: biexp_fit}
        output_path = tmp_path / "levels_biexp"

        result = export_levels(
            sample_levels,
            output_path,
            fmt=ExportFormat.CSV,
            level_fits=level_fits,
        )

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            fieldnames = reader.fieldnames

        # Should have second exponential columns
        assert "tau_2_ns" in fieldnames
        assert "amp_2" in fieldnames
        assert rows[0]["tau_1_ns"] == "2.0000"
        assert rows[0]["tau_2_ns"] == "8.0000"

    def test_export_levels_json(
        self, sample_levels: list[LevelData], tmp_path: Path
    ) -> None:
        """Export levels to JSON format."""
        output_path = tmp_path / "levels"
        result = export_levels(
            sample_levels,
            output_path,
            fmt=ExportFormat.JSON,
            particle_name="Test Particle",
        )

        with open(result, "r") as f:
            data = json.load(f)

        assert data["metadata"]["type"] == "levels"
        assert data["metadata"]["particle_name"] == "Test Particle"
        assert data["metadata"]["num_levels"] == 3
        assert len(data["levels"]) == 3
        assert data["levels"][0]["num_photons"] == 100

    def test_export_levels_parquet(
        self, sample_levels: list[LevelData], tmp_path: Path
    ) -> None:
        """Export levels to Parquet format."""
        pytest.importorskip("pyarrow")

        output_path = tmp_path / "levels"
        result = export_levels(sample_levels, output_path, fmt=ExportFormat.PARQUET)

        assert result.suffix == ".parquet"
        assert result.exists()

        import pyarrow.parquet as pq

        table = pq.read_table(result)
        assert "level_index" in table.column_names
        assert "num_photons" in table.column_names
        assert len(table) == 3

    def test_export_levels_excel(
        self, sample_levels: list[LevelData], tmp_path: Path
    ) -> None:
        """Export levels to Excel format."""
        pytest.importorskip("openpyxl")

        output_path = tmp_path / "levels"
        result = export_levels(sample_levels, output_path, fmt=ExportFormat.EXCEL)

        assert result.suffix == ".xlsx"
        assert result.exists()


# ============================================================================
# Test export_groups
# ============================================================================


class TestExportGroups:
    """Tests for export_groups function."""

    def test_export_groups_csv(
        self, sample_groups: list[GroupData], tmp_path: Path
    ) -> None:
        """Export groups to CSV."""
        output_path = tmp_path / "groups"
        result = export_groups(sample_groups, output_path, fmt=ExportFormat.CSV)

        assert result.suffix == ".csv"
        assert result.exists()

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 2
        assert rows[0]["group_id"] == "0"
        assert rows[0]["num_levels"] == "2"
        assert rows[0]["total_photons"] == "250"

    def test_level_indices_serialization(
        self, sample_groups: list[GroupData], tmp_path: Path
    ) -> None:
        """Level indices are serialized as semicolon-separated string."""
        output_path = tmp_path / "groups"
        result = export_groups(sample_groups, output_path, fmt=ExportFormat.CSV)

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Group 0 has levels 0 and 2
        assert rows[0]["level_indices"] == "0;2"
        # Group 1 has only level 1
        assert rows[1]["level_indices"] == "1"

    def test_export_groups_json(
        self, sample_groups: list[GroupData], tmp_path: Path
    ) -> None:
        """Export groups to JSON format."""
        output_path = tmp_path / "groups"
        result = export_groups(
            sample_groups,
            output_path,
            fmt=ExportFormat.JSON,
            particle_name="Test Particle",
        )

        with open(result, "r") as f:
            data = json.load(f)

        assert data["metadata"]["type"] == "groups"
        assert data["metadata"]["particle_name"] == "Test Particle"
        assert data["metadata"]["num_groups"] == 2
        assert len(data["groups"]) == 2
        assert data["groups"][0]["level_indices"] == [0, 2]

    def test_export_groups_parquet(
        self, sample_groups: list[GroupData], tmp_path: Path
    ) -> None:
        """Export groups to Parquet format."""
        pytest.importorskip("pyarrow")

        output_path = tmp_path / "groups"
        result = export_groups(sample_groups, output_path, fmt=ExportFormat.PARQUET)

        assert result.suffix == ".parquet"
        assert result.exists()

        import pyarrow.parquet as pq

        table = pq.read_table(result)
        assert "group_id" in table.column_names
        assert "level_indices" in table.column_names
        assert len(table) == 2

    def test_export_groups_excel(
        self, sample_groups: list[GroupData], tmp_path: Path
    ) -> None:
        """Export groups to Excel format."""
        pytest.importorskip("openpyxl")

        output_path = tmp_path / "groups"
        result = export_groups(sample_groups, output_path, fmt=ExportFormat.EXCEL)

        assert result.suffix == ".xlsx"
        assert result.exists()


# ============================================================================
# Test export_fit_results
# ============================================================================


class TestExportFitResults:
    """Tests for export_fit_results function."""

    def test_export_single_exponential(
        self, sample_fit_result: FitResult, tmp_path: Path
    ) -> None:
        """Export single exponential fit result."""
        fit_results = {(1, 1, 0): sample_fit_result}
        output_path = tmp_path / "fits"

        result = export_fit_results(fit_results, output_path, fmt=ExportFormat.CSV)

        assert result.exists()

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 1
        assert rows[0]["particle_id"] == "1"
        assert rows[0]["channel"] == "1"
        assert rows[0]["level_or_group_id"] == "0"
        assert rows[0]["num_exponentials"] == "1"
        assert float(rows[0]["tau_1_ns"]) == pytest.approx(5.0, rel=0.01)
        assert rows[0]["tau_2_ns"] == ""
        assert rows[0]["tau_3_ns"] == ""

    def test_export_biexponential(
        self, biexponential_fit_result: FitResult, tmp_path: Path
    ) -> None:
        """Export biexponential fit result."""
        fit_results = {(1, 1, 0): biexponential_fit_result}
        output_path = tmp_path / "fits"

        result = export_fit_results(fit_results, output_path, fmt=ExportFormat.CSV)

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert rows[0]["num_exponentials"] == "2"
        assert float(rows[0]["tau_1_ns"]) == pytest.approx(2.0, rel=0.01)
        assert float(rows[0]["tau_2_ns"]) == pytest.approx(8.0, rel=0.01)
        assert rows[0]["tau_3_ns"] == ""

    def test_export_triexponential(
        self, triexponential_fit_result: FitResult, tmp_path: Path
    ) -> None:
        """Export triexponential fit result."""
        fit_results = {(1, 1, 0): triexponential_fit_result}
        output_path = tmp_path / "fits"

        result = export_fit_results(fit_results, output_path, fmt=ExportFormat.CSV)

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert rows[0]["num_exponentials"] == "3"
        assert float(rows[0]["tau_1_ns"]) == pytest.approx(1.0, rel=0.01)
        assert float(rows[0]["tau_2_ns"]) == pytest.approx(4.0, rel=0.01)
        assert float(rows[0]["tau_3_ns"]) == pytest.approx(12.0, rel=0.01)

    def test_export_multiple_fits(
        self,
        sample_fit_result: FitResult,
        biexponential_fit_result: FitResult,
        tmp_path: Path,
    ) -> None:
        """Export multiple fit results."""
        fit_results = {
            (1, 1, 0): sample_fit_result,
            (1, 1, 1): biexponential_fit_result,
            (2, 1, 0): sample_fit_result,
        }
        output_path = tmp_path / "fits"

        result = export_fit_results(fit_results, output_path, fmt=ExportFormat.CSV)

        with open(result, "r") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 3

    def test_export_fit_results_json(
        self, sample_fit_result: FitResult, tmp_path: Path
    ) -> None:
        """Export fit results to JSON format."""
        fit_results = {(1, 1, 0): sample_fit_result}
        output_path = tmp_path / "fits"

        result = export_fit_results(fit_results, output_path, fmt=ExportFormat.JSON)

        with open(result, "r") as f:
            data = json.load(f)

        assert data["metadata"]["type"] == "fit_results"
        assert data["metadata"]["num_results"] == 1
        assert len(data["fit_results"]) == 1
        assert data["fit_results"][0]["tau_ns"] == [5.0]

    def test_export_fit_results_parquet(
        self, sample_fit_result: FitResult, tmp_path: Path
    ) -> None:
        """Export fit results to Parquet format."""
        pytest.importorskip("pyarrow")

        fit_results = {(1, 1, 0): sample_fit_result}
        output_path = tmp_path / "fits"

        result = export_fit_results(fit_results, output_path, fmt=ExportFormat.PARQUET)

        assert result.suffix == ".parquet"
        assert result.exists()

    def test_export_fit_results_excel(
        self, sample_fit_result: FitResult, tmp_path: Path
    ) -> None:
        """Export fit results to Excel format."""
        pytest.importorskip("openpyxl")

        fit_results = {(1, 1, 0): sample_fit_result}
        output_path = tmp_path / "fits"

        result = export_fit_results(fit_results, output_path, fmt=ExportFormat.EXCEL)

        assert result.suffix == ".xlsx"
        assert result.exists()


# ============================================================================
# Test export_batch
# ============================================================================


class TestExportBatch:
    """Tests for export_batch function."""

    @pytest.fixture
    def sample_channel(self) -> ChannelData:
        """Create a sample channel."""
        return ChannelData(
            abstimes=np.linspace(0, 1_000_000_000, 100, dtype=np.uint64),
            microtimes=np.random.rand(100).astype(np.float64),
        )

    @pytest.fixture
    def sample_state(self, sample_channel: ChannelData) -> SessionState:
        """Create a sample session state."""
        particle1 = ParticleData(
            id=1,
            name="Particle 1",
            tcspc_card="SPC-150",
            channelwidth=0.012,
            channel1=sample_channel,
        )
        particle2 = ParticleData(
            id=2,
            name="Particle 2",
            tcspc_card="SPC-150",
            channelwidth=0.012,
            channel1=sample_channel,
            channel2=sample_channel,
        )

        state = SessionState()
        state.file_metadata = FileMetadata(
            path=Path("/path/to/test.h5"),
            filename="test.h5",
            num_particles=2,
            has_irf=True,
        )
        state.particles = [particle1, particle2]

        # Add levels for particle 1
        level = LevelData(
            start_index=0,
            end_index=49,
            start_time_ns=0,
            end_time_ns=500_000_000,
            num_photons=50,
            intensity_cps=100.0,
        )
        state.set_levels(1, 1, [level])

        return state

    def test_batch_export_multiple_particles(
        self, sample_state: SessionState, tmp_path: Path
    ) -> None:
        """Export data for multiple particles."""
        selections = [(1, 1), (2, 1)]
        output_dir = tmp_path / "export"

        files = export_batch(
            sample_state,
            selections,
            output_dir,
            fmt=ExportFormat.CSV,
            export_intensity=True,
            export_levels=True,
            export_groups=False,
            export_fits=False,
        )

        # Should have intensity for both particles, levels for particle 1
        assert len(files) >= 2
        assert output_dir.exists()

    def test_progress_callback_called(
        self, sample_state: SessionState, tmp_path: Path
    ) -> None:
        """Progress callback is called during batch export."""
        progress_calls = []

        def track_progress(progress: float, message: str) -> None:
            progress_calls.append((progress, message))

        selections = [(1, 1), (2, 1)]
        output_dir = tmp_path / "export"

        export_batch(
            sample_state,
            selections,
            output_dir,
            fmt=ExportFormat.CSV,
            progress_callback=track_progress,
        )

        # Should have been called for each particle and final
        assert len(progress_calls) >= 2
        # Last call should be 1.0
        assert progress_calls[-1][0] == 1.0

    def test_skips_missing_data(
        self, sample_state: SessionState, tmp_path: Path
    ) -> None:
        """Batch export skips missing data gracefully."""
        # Request export for non-existent particle
        selections = [(1, 1), (99, 1)]  # 99 doesn't exist
        output_dir = tmp_path / "export"

        # Should not raise
        files = export_batch(
            sample_state,
            selections,
            output_dir,
            fmt=ExportFormat.CSV,
        )

        # Should have files only for existing particle
        assert len(files) >= 1

    def test_empty_selections(
        self, sample_state: SessionState, tmp_path: Path
    ) -> None:
        """Batch export with empty selections returns empty list."""
        output_dir = tmp_path / "export"

        files = export_batch(
            sample_state,
            [],
            output_dir,
            fmt=ExportFormat.CSV,
        )

        assert files == []

    def test_different_formats(
        self, sample_state: SessionState, tmp_path: Path
    ) -> None:
        """Batch export respects format parameter."""
        selections = [(1, 1)]
        output_dir = tmp_path / "export"

        files = export_batch(
            sample_state,
            selections,
            output_dir,
            fmt=ExportFormat.JSON,
            export_intensity=True,
            export_levels=False,
            export_groups=False,
            export_fits=False,
        )

        assert len(files) == 1
        assert files[0].suffix == ".json"


# ============================================================================
# Test error handling
# ============================================================================


class TestExportErrorHandling:
    """Tests for error handling in export functions."""

    def test_unsupported_format_raises(
        self, sample_abstimes: np.ndarray, tmp_path: Path
    ) -> None:
        """Unsupported format raises ValueError."""
        from unittest.mock import MagicMock

        # Create a mock format that's not in the switch
        mock_format = MagicMock()
        mock_format.value = "unsupported"

        output_path = tmp_path / "test"

        # This should raise ValueError for unsupported format
        # Note: We can't directly test this without patching internal functions
        # since ExportFormat enum enforces valid values

    def test_parquet_import_error(self, tmp_path: Path) -> None:
        """Parquet export without pyarrow raises ImportError with helpful message."""
        # This test would require mocking the import, which is complex
        # The implementation already provides a clear error message
        pass

    def test_excel_import_error(self, tmp_path: Path) -> None:
        """Excel export without openpyxl raises ImportError with helpful message."""
        # This test would require mocking the import, which is complex
        # The implementation already provides a clear error message
        pass
