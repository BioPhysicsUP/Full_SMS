"""Data export functions for Full SMS analysis results.

This module provides functions to export intensity traces, levels, groups,
and fit results to various file formats (CSV, Parquet, Excel).
"""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import asdict
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np
from numpy.typing import NDArray

from full_sms.analysis.histograms import bin_photons

if TYPE_CHECKING:
    from full_sms.models.fit import FitResult
    from full_sms.models.group import ClusteringResult, GroupData
    from full_sms.models.level import LevelData
    from full_sms.models.measurement import MeasurementData
    from full_sms.models.session import SessionState


logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """Supported export formats."""

    CSV = "csv"
    PARQUET = "parquet"
    EXCEL = "xlsx"
    JSON = "json"


def _get_extension(fmt: ExportFormat) -> str:
    """Get the file extension for a format."""
    return f".{fmt.value}"


def _ensure_directory(path: Path) -> None:
    """Ensure the parent directory exists."""
    path.parent.mkdir(parents=True, exist_ok=True)


def export_intensity_trace(
    abstimes: NDArray[np.uint64],
    output_path: Path,
    bin_size_ms: float = 10.0,
    fmt: ExportFormat = ExportFormat.CSV,
    measurement_name: str = "",
) -> Path:
    """Export binned intensity trace to a file.

    Args:
        abstimes: Absolute photon arrival times in nanoseconds.
        output_path: Path to output file (without extension).
        bin_size_ms: Bin size in milliseconds.
        fmt: Export format.
        measurement_name: Optional measurement name for metadata.

    Returns:
        Path to the exported file.
    """
    # Bin the photons
    times_ms, counts = bin_photons(abstimes.astype(np.float64), bin_size_ms)

    # Calculate intensity in cps
    bin_size_s = bin_size_ms / 1000.0
    intensity_cps = counts.astype(np.float64) / bin_size_s

    # Build output path with extension
    output_file = output_path.with_suffix(_get_extension(fmt))
    _ensure_directory(output_file)

    if fmt == ExportFormat.CSV:
        _export_intensity_csv(output_file, times_ms, counts, intensity_cps)
    elif fmt == ExportFormat.PARQUET:
        _export_intensity_parquet(output_file, times_ms, counts, intensity_cps)
    elif fmt == ExportFormat.EXCEL:
        _export_intensity_excel(output_file, times_ms, counts, intensity_cps)
    elif fmt == ExportFormat.JSON:
        _export_intensity_json(
            output_file, times_ms, counts, intensity_cps, bin_size_ms, measurement_name
        )
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    logger.info(f"Exported intensity trace to {output_file}")
    return output_file


def _export_intensity_csv(
    path: Path,
    times_ms: NDArray[np.float64],
    counts: NDArray[np.int64],
    intensity_cps: NDArray[np.float64],
) -> None:
    """Export intensity trace to CSV."""
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["time_ms", "counts", "intensity_cps"])
        for t, c, i in zip(times_ms, counts, intensity_cps):
            writer.writerow([f"{t:.3f}", c, f"{i:.2f}"])


def _export_intensity_parquet(
    path: Path,
    times_ms: NDArray[np.float64],
    counts: NDArray[np.int64],
    intensity_cps: NDArray[np.float64],
) -> None:
    """Export intensity trace to Parquet format."""
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError("pyarrow is required for Parquet export. Install with: pip install pyarrow")

    table = pa.table({
        "time_ms": times_ms,
        "counts": counts,
        "intensity_cps": intensity_cps,
    })
    pq.write_table(table, path)


def _export_intensity_excel(
    path: Path,
    times_ms: NDArray[np.float64],
    counts: NDArray[np.int64],
    intensity_cps: NDArray[np.float64],
) -> None:
    """Export intensity trace to Excel format."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Intensity Trace"

    # Header
    ws.append(["time_ms", "counts", "intensity_cps"])

    # Data
    for t, c, i in zip(times_ms, counts, intensity_cps):
        ws.append([float(t), int(c), float(i)])

    wb.save(path)


def _export_intensity_json(
    path: Path,
    times_ms: NDArray[np.float64],
    counts: NDArray[np.int64],
    intensity_cps: NDArray[np.float64],
    bin_size_ms: float,
    measurement_name: str,
) -> None:
    """Export intensity trace to JSON format."""
    data = {
        "metadata": {
            "type": "intensity_trace",
            "bin_size_ms": bin_size_ms,
            "measurement_name": measurement_name,
            "num_bins": len(times_ms),
        },
        "data": {
            "time_ms": times_ms.tolist(),
            "counts": counts.tolist(),
            "intensity_cps": intensity_cps.tolist(),
        },
    }
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def export_levels(
    levels: list[LevelData],
    output_path: Path,
    fmt: ExportFormat = ExportFormat.CSV,
    measurement_name: str = "",
    level_fits: dict[int, "FitResultData"] | None = None,
) -> Path:
    """Export detected levels to a file.

    Args:
        levels: List of LevelData from change point analysis.
        output_path: Path to output file (without extension).
        fmt: Export format.
        measurement_name: Optional measurement name for metadata.
        level_fits: Optional dict mapping level_index to FitResultData for that level.

    Returns:
        Path to the exported file.
    """
    output_file = output_path.with_suffix(_get_extension(fmt))
    _ensure_directory(output_file)

    if fmt == ExportFormat.CSV:
        _export_levels_csv(output_file, levels, level_fits)
    elif fmt == ExportFormat.PARQUET:
        _export_levels_parquet(output_file, levels, level_fits)
    elif fmt == ExportFormat.EXCEL:
        _export_levels_excel(output_file, levels, level_fits)
    elif fmt == ExportFormat.JSON:
        _export_levels_json(output_file, levels, measurement_name, level_fits)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    logger.info(f"Exported {len(levels)} levels to {output_file}")
    return output_file


def _export_levels_csv(path: Path, levels: list[LevelData], level_fits: dict[int, "FitResultData"] | None = None) -> None:
    """Export levels to CSV with optional fit parameters."""
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)

        # Build header row
        header = [
            "level_index",
            "start_time_s",
            "end_time_s",
            "dwell_time_s",
            "num_photons",
            "intensity_cps",
            "group_id",
        ]

        # Determine maximum number of exponentials across all fits
        max_exponentials = 0
        if level_fits:
            for fit in level_fits.values():
                max_exponentials = max(max_exponentials, fit.num_exponentials)

        # Add fit parameter columns based on max_exponentials
        if level_fits and max_exponentials > 0:
            # Always add first exponential
            header.extend(["tau_1_ns", "tau_1_std", "amp_1"])

            # Add second exponential if needed
            if max_exponentials >= 2:
                header.extend(["tau_2_ns", "tau_2_std", "amp_2"])

            # Add third exponential if needed
            if max_exponentials >= 3:
                header.extend(["tau_3_ns", "tau_3_std", "amp_3"])

            # Always add summary statistics
            header.extend(["avg_lifetime_ns", "chi_squared", "durbin_watson"])

        writer.writerow(header)

        for idx, level in enumerate(levels):
            row = [
                idx,
                f"{level.start_time_ns / 1e9:.6f}",
                f"{level.end_time_ns / 1e9:.6f}",
                f"{level.dwell_time_s:.6f}",
                level.num_photons,
                f"{level.intensity_cps:.2f}",
                level.group_id if level.group_id is not None else "",
            ]

            # Add fit parameters if available for this level
            if level_fits and idx in level_fits:
                fit = level_fits[idx]

                # First exponential
                row.extend([
                    f"{fit.tau[0]:.4f}" if len(fit.tau) > 0 else "",
                    f"{fit.tau_std[0]:.4f}" if len(fit.tau_std) > 0 else "",
                    f"{fit.amplitude[0]:.4f}" if len(fit.amplitude) > 0 else "",
                ])

                # Second exponential (if columns exist)
                if max_exponentials >= 2:
                    row.extend([
                        f"{fit.tau[1]:.4f}" if len(fit.tau) > 1 else "",
                        f"{fit.tau_std[1]:.4f}" if len(fit.tau_std) > 1 else "",
                        f"{fit.amplitude[1]:.4f}" if len(fit.amplitude) > 1 else "",
                    ])

                # Third exponential (if columns exist)
                if max_exponentials >= 3:
                    row.extend([
                        f"{fit.tau[2]:.4f}" if len(fit.tau) > 2 else "",
                        f"{fit.tau_std[2]:.4f}" if len(fit.tau_std) > 2 else "",
                        f"{fit.amplitude[2]:.4f}" if len(fit.amplitude) > 2 else "",
                    ])

                # Summary statistics
                row.extend([
                    f"{fit.average_lifetime:.4f}",
                    f"{fit.chi_squared:.4f}",
                    f"{fit.durbin_watson:.4f}",
                ])
            elif level_fits:
                # Add empty values if fits are expected but not available for this level
                num_empty = 3 + (3 if max_exponentials >= 2 else 0) + (3 if max_exponentials >= 3 else 0) + 3
                row.extend([""] * num_empty)

            writer.writerow(row)


def _export_levels_parquet(path: Path, levels: list[LevelData], level_fits: dict[int, "FitResultData"] | None = None) -> None:
    """Export levels to Parquet format."""
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError("pyarrow is required for Parquet export. Install with: pip install pyarrow")

    data = {
        "level_index": list(range(len(levels))),
        "start_time_s": [level.start_time_ns / 1e9 for level in levels],
        "end_time_s": [level.end_time_ns / 1e9 for level in levels],
        "dwell_time_s": [level.dwell_time_s for level in levels],
        "num_photons": [level.num_photons for level in levels],
        "intensity_cps": [level.intensity_cps for level in levels],
        "group_id": [level.group_id for level in levels],
    }

    # Determine maximum number of exponentials
    max_exponentials = 0
    if level_fits:
        for fit in level_fits.values():
            max_exponentials = max(max_exponentials, fit.num_exponentials)

    # Add fit parameters based on max_exponentials
    if level_fits and max_exponentials > 0:
        # First exponential (always present)
        data.update({
            "tau_1_ns": [level_fits[i].tau[0] if i in level_fits and len(level_fits[i].tau) > 0 else None for i in range(len(levels))],
            "tau_1_std": [level_fits[i].tau_std[0] if i in level_fits and len(level_fits[i].tau_std) > 0 else None for i in range(len(levels))],
            "amp_1": [level_fits[i].amplitude[0] if i in level_fits and len(level_fits[i].amplitude) > 0 else None for i in range(len(levels))],
        })

        # Second exponential (if used)
        if max_exponentials >= 2:
            data.update({
                "tau_2_ns": [level_fits[i].tau[1] if i in level_fits and len(level_fits[i].tau) > 1 else None for i in range(len(levels))],
                "tau_2_std": [level_fits[i].tau_std[1] if i in level_fits and len(level_fits[i].tau_std) > 1 else None for i in range(len(levels))],
                "amp_2": [level_fits[i].amplitude[1] if i in level_fits and len(level_fits[i].amplitude) > 1 else None for i in range(len(levels))],
            })

        # Third exponential (if used)
        if max_exponentials >= 3:
            data.update({
                "tau_3_ns": [level_fits[i].tau[2] if i in level_fits and len(level_fits[i].tau) > 2 else None for i in range(len(levels))],
                "tau_3_std": [level_fits[i].tau_std[2] if i in level_fits and len(level_fits[i].tau_std) > 2 else None for i in range(len(levels))],
                "amp_3": [level_fits[i].amplitude[2] if i in level_fits and len(level_fits[i].amplitude) > 2 else None for i in range(len(levels))],
            })

        # Summary statistics (always present when fits exist)
        data.update({
            "avg_lifetime_ns": [level_fits[i].average_lifetime if i in level_fits else None for i in range(len(levels))],
            "chi_squared": [level_fits[i].chi_squared if i in level_fits else None for i in range(len(levels))],
            "durbin_watson": [level_fits[i].durbin_watson if i in level_fits else None for i in range(len(levels))],
        })

    table = pa.table(data)
    pq.write_table(table, path)


def _export_levels_excel(path: Path, levels: list[LevelData], level_fits: dict[int, "FitResultData"] | None = None) -> None:
    """Export levels to Excel format."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Levels"

    # Header
    ws.append([
        "level_index",
        "start_time_s",
        "end_time_s",
        "dwell_time_s",
        "num_photons",
        "intensity_cps",
        "group_id",
    ])

    # Data
    for idx, level in enumerate(levels):
        ws.append([
            idx,
            level.start_time_ns / 1e9,
            level.end_time_ns / 1e9,
            level.dwell_time_s,
            level.num_photons,
            level.intensity_cps,
            level.group_id,
        ])

    wb.save(path)


def _export_levels_json(path: Path, levels: list[LevelData], measurement_name: str, level_fits: dict[int, "FitResultData"] | None = None) -> None:
    """Export levels to JSON format."""
    data = {
        "metadata": {
            "type": "levels",
            "measurement_name": measurement_name,
            "num_levels": len(levels),
        },
        "levels": [
            {
                "index": idx,
                "start_time_s": level.start_time_ns / 1e9,
                "end_time_s": level.end_time_ns / 1e9,
                "dwell_time_s": level.dwell_time_s,
                "num_photons": level.num_photons,
                "intensity_cps": level.intensity_cps,
                "group_id": level.group_id,
            }
            for idx, level in enumerate(levels)
        ],
    }
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def export_groups(
    groups: list[GroupData],
    output_path: Path,
    fmt: ExportFormat = ExportFormat.CSV,
    measurement_name: str = "",
) -> Path:
    """Export clustered groups to a file.

    Args:
        groups: List of GroupData from clustering.
        output_path: Path to output file (without extension).
        fmt: Export format.
        measurement_name: Optional measurement name for metadata.

    Returns:
        Path to the exported file.
    """
    output_file = output_path.with_suffix(_get_extension(fmt))
    _ensure_directory(output_file)

    if fmt == ExportFormat.CSV:
        _export_groups_csv(output_file, groups)
    elif fmt == ExportFormat.PARQUET:
        _export_groups_parquet(output_file, groups)
    elif fmt == ExportFormat.EXCEL:
        _export_groups_excel(output_file, groups)
    elif fmt == ExportFormat.JSON:
        _export_groups_json(output_file, groups, measurement_name)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    logger.info(f"Exported {len(groups)} groups to {output_file}")
    return output_file


def _export_groups_csv(path: Path, groups: list[GroupData]) -> None:
    """Export groups to CSV."""
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "group_id",
            "num_levels",
            "total_photons",
            "total_dwell_time_s",
            "intensity_cps",
            "level_indices",
        ])
        for group in groups:
            writer.writerow([
                group.group_id,
                group.num_levels,
                group.total_photons,
                f"{group.total_dwell_time_s:.6f}",
                f"{group.intensity_cps:.2f}",
                ";".join(str(i) for i in group.level_indices),
            ])


def _export_groups_parquet(path: Path, groups: list[GroupData]) -> None:
    """Export groups to Parquet format."""
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError("pyarrow is required for Parquet export. Install with: pip install pyarrow")

    data = {
        "group_id": [group.group_id for group in groups],
        "num_levels": [group.num_levels for group in groups],
        "total_photons": [group.total_photons for group in groups],
        "total_dwell_time_s": [group.total_dwell_time_s for group in groups],
        "intensity_cps": [group.intensity_cps for group in groups],
        "level_indices": [list(group.level_indices) for group in groups],
    }
    table = pa.table(data)
    pq.write_table(table, path)


def _export_groups_excel(path: Path, groups: list[GroupData]) -> None:
    """Export groups to Excel format."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Groups"

    # Header
    ws.append([
        "group_id",
        "num_levels",
        "total_photons",
        "total_dwell_time_s",
        "intensity_cps",
        "level_indices",
    ])

    # Data
    for group in groups:
        ws.append([
            group.group_id,
            group.num_levels,
            group.total_photons,
            group.total_dwell_time_s,
            group.intensity_cps,
            ";".join(str(i) for i in group.level_indices),
        ])

    wb.save(path)


def _export_groups_json(path: Path, groups: list[GroupData], measurement_name: str) -> None:
    """Export groups to JSON format."""
    data = {
        "metadata": {
            "type": "groups",
            "measurement_name": measurement_name,
            "num_groups": len(groups),
        },
        "groups": [
            {
                "group_id": group.group_id,
                "num_levels": group.num_levels,
                "total_photons": group.total_photons,
                "total_dwell_time_s": group.total_dwell_time_s,
                "intensity_cps": group.intensity_cps,
                "level_indices": list(group.level_indices),
            }
            for group in groups
        ],
    }
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def export_fit_results(
    fit_results: dict[tuple[int, int, int], FitResult],
    output_path: Path,
    fmt: ExportFormat = ExportFormat.CSV,
) -> Path:
    """Export lifetime fit results to a file.

    Args:
        fit_results: Dict mapping (measurement_id, channel, level_or_group_id) to FitResult.
        output_path: Path to output file (without extension).
        fmt: Export format.

    Returns:
        Path to the exported file.
    """
    output_file = output_path.with_suffix(_get_extension(fmt))
    _ensure_directory(output_file)

    if fmt == ExportFormat.CSV:
        _export_fit_results_csv(output_file, fit_results)
    elif fmt == ExportFormat.PARQUET:
        _export_fit_results_parquet(output_file, fit_results)
    elif fmt == ExportFormat.EXCEL:
        _export_fit_results_excel(output_file, fit_results)
    elif fmt == ExportFormat.JSON:
        _export_fit_results_json(output_file, fit_results)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    logger.info(f"Exported {len(fit_results)} fit results to {output_file}")
    return output_file


def _export_fit_results_csv(
    path: Path, fit_results: dict[tuple[int, int, int], FitResult]
) -> None:
    """Export fit results to CSV."""
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "measurement_id",
            "channel",
            "level_or_group_id",
            "num_exponentials",
            "tau_1_ns",
            "tau_1_std",
            "amp_1",
            "tau_2_ns",
            "tau_2_std",
            "amp_2",
            "tau_3_ns",
            "tau_3_std",
            "amp_3",
            "avg_lifetime_ns",
            "chi_squared",
            "durbin_watson",
            "shift",
            "background",
        ])
        for (mid, ch, level_id), fit in fit_results.items():
            row = [
                mid,
                ch,
                level_id,
                fit.num_exponentials,
                f"{fit.tau[0]:.4f}" if len(fit.tau) > 0 else "",
                f"{fit.tau_std[0]:.4f}" if len(fit.tau_std) > 0 else "",
                f"{fit.amplitude[0]:.4f}" if len(fit.amplitude) > 0 else "",
                f"{fit.tau[1]:.4f}" if len(fit.tau) > 1 else "",
                f"{fit.tau_std[1]:.4f}" if len(fit.tau_std) > 1 else "",
                f"{fit.amplitude[1]:.4f}" if len(fit.amplitude) > 1 else "",
                f"{fit.tau[2]:.4f}" if len(fit.tau) > 2 else "",
                f"{fit.tau_std[2]:.4f}" if len(fit.tau_std) > 2 else "",
                f"{fit.amplitude[2]:.4f}" if len(fit.amplitude) > 2 else "",
                f"{fit.average_lifetime:.4f}",
                f"{fit.chi_squared:.4f}",
                f"{fit.durbin_watson:.4f}",
                f"{fit.shift:.4f}",
                f"{fit.background:.4f}",
            ]
            writer.writerow(row)


def _export_fit_results_parquet(
    path: Path, fit_results: dict[tuple[int, int, int], FitResult]
) -> None:
    """Export fit results to Parquet format."""
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError("pyarrow is required for Parquet export. Install with: pip install pyarrow")

    data: dict[str, list] = {
        "measurement_id": [],
        "channel": [],
        "level_or_group_id": [],
        "num_exponentials": [],
        "tau_1_ns": [],
        "tau_2_ns": [],
        "tau_3_ns": [],
        "amp_1": [],
        "amp_2": [],
        "amp_3": [],
        "avg_lifetime_ns": [],
        "chi_squared": [],
        "durbin_watson": [],
        "shift": [],
        "background": [],
    }

    for (mid, ch, level_id), fit in fit_results.items():
        data["measurement_id"].append(mid)
        data["channel"].append(ch)
        data["level_or_group_id"].append(level_id)
        data["num_exponentials"].append(fit.num_exponentials)
        data["tau_1_ns"].append(fit.tau[0] if len(fit.tau) > 0 else None)
        data["tau_2_ns"].append(fit.tau[1] if len(fit.tau) > 1 else None)
        data["tau_3_ns"].append(fit.tau[2] if len(fit.tau) > 2 else None)
        data["amp_1"].append(fit.amplitude[0] if len(fit.amplitude) > 0 else None)
        data["amp_2"].append(fit.amplitude[1] if len(fit.amplitude) > 1 else None)
        data["amp_3"].append(fit.amplitude[2] if len(fit.amplitude) > 2 else None)
        data["avg_lifetime_ns"].append(fit.average_lifetime)
        data["chi_squared"].append(fit.chi_squared)
        data["durbin_watson"].append(fit.durbin_watson)
        data["shift"].append(fit.shift)
        data["background"].append(fit.background)

    table = pa.table(data)
    pq.write_table(table, path)


def _export_fit_results_excel(
    path: Path, fit_results: dict[tuple[int, int, int], FitResult]
) -> None:
    """Export fit results to Excel format."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fit Results"

    # Header
    ws.append([
        "measurement_id",
        "channel",
        "level_or_group_id",
        "num_exponentials",
        "tau_1_ns",
        "tau_1_std",
        "amp_1",
        "tau_2_ns",
        "tau_2_std",
        "amp_2",
        "tau_3_ns",
        "tau_3_std",
        "amp_3",
        "avg_lifetime_ns",
        "chi_squared",
        "durbin_watson",
        "shift",
        "background",
    ])

    # Data
    for (mid, ch, level_id), fit in fit_results.items():
        ws.append([
            mid,
            ch,
            level_id,
            fit.num_exponentials,
            fit.tau[0] if len(fit.tau) > 0 else None,
            fit.tau_std[0] if len(fit.tau_std) > 0 else None,
            fit.amplitude[0] if len(fit.amplitude) > 0 else None,
            fit.tau[1] if len(fit.tau) > 1 else None,
            fit.tau_std[1] if len(fit.tau_std) > 1 else None,
            fit.amplitude[1] if len(fit.amplitude) > 1 else None,
            fit.tau[2] if len(fit.tau) > 2 else None,
            fit.tau_std[2] if len(fit.tau_std) > 2 else None,
            fit.amplitude[2] if len(fit.amplitude) > 2 else None,
            fit.average_lifetime,
            fit.chi_squared,
            fit.durbin_watson,
            fit.shift,
            fit.background,
        ])

    wb.save(path)


def _export_fit_results_json(
    path: Path, fit_results: dict[tuple[int, int, int], FitResult]
) -> None:
    """Export fit results to JSON format."""
    results = []
    for (mid, ch, level_id), fit in fit_results.items():
        results.append({
            "measurement_id": mid,
            "channel": ch,
            "level_or_group_id": level_id,
            "num_exponentials": fit.num_exponentials,
            "tau_ns": list(fit.tau),
            "tau_std_ns": list(fit.tau_std),
            "amplitude": list(fit.amplitude),
            "amplitude_std": list(fit.amplitude_std),
            "average_lifetime_ns": fit.average_lifetime,
            "chi_squared": fit.chi_squared,
            "durbin_watson": fit.durbin_watson,
            "shift": fit.shift,
            "background": fit.background,
        })

    data = {
        "metadata": {
            "type": "fit_results",
            "num_results": len(results),
        },
        "fit_results": results,
    }
    with open(path, "w") as f:
        json.dump(data, f, indent=2)


def export_all_measurement_data(
    state: SessionState,
    measurement_id: int,
    channel: int,
    output_dir: Path,
    fmt: ExportFormat = ExportFormat.CSV,
    export_intensity: bool = True,
    export_levels: bool = True,
    export_groups: bool = True,
    export_fits: bool = True,
    bin_size_ms: float = 10.0,
) -> list[Path]:
    """Export all analysis data for a single measurement.

    Args:
        state: The session state containing all data.
        measurement_id: The measurement ID to export.
        channel: The channel number (1 or 2).
        output_dir: Directory to export files to.
        fmt: Export format.
        export_intensity: Whether to export intensity trace.
        export_levels: Whether to export levels.
        export_groups: Whether to export groups.
        export_fits: Whether to export fit results.
        bin_size_ms: Bin size for intensity trace.

    Returns:
        List of paths to exported files.
    """
    exported_files: list[Path] = []
    measurement = state.get_measurement(measurement_id)

    if measurement is None:
        logger.warning(f"Measurement {measurement_id} not found")
        return exported_files

    prefix = f"measurement_{measurement_id}_ch{channel}"

    # Export intensity trace
    if export_intensity:
        channel_data = measurement.channel1 if channel == 1 else measurement.channel2
        if channel_data is not None:
            path = export_intensity_trace(
                channel_data.abstimes,
                output_dir / f"{prefix}_intensity",
                bin_size_ms=bin_size_ms,
                fmt=fmt,
                measurement_name=measurement.name,
            )
            exported_files.append(path)

    # Export levels
    if export_levels:
        levels_data = state.get_levels(measurement_id, channel)
        if levels_data:
            # Get level-specific fit results if available
            level_fits_dict = {}
            all_level_fits = state.get_all_level_fits(measurement_id, channel)
            if all_level_fits:
                for level_index, fit_data in all_level_fits.items():
                    level_fits_dict[level_index] = fit_data

            from full_sms.io.exporters import export_levels as _export_levels
            path = _export_levels(
                levels_data,
                output_dir / f"{prefix}_levels",
                fmt=fmt,
                measurement_name=measurement.name,
                level_fits=level_fits_dict if level_fits_dict else None,
            )
            exported_files.append(path)

    # Export groups
    if export_groups:
        groups_data = state.get_groups(measurement_id, channel)
        if groups_data:
            from full_sms.io.exporters import export_groups as _export_groups
            path = _export_groups(
                groups_data,
                output_dir / f"{prefix}_groups",
                fmt=fmt,
                measurement_name=measurement.name,
            )
            exported_files.append(path)

    # Export fit results for this measurement
    # Note: Fit export requires FitResult (with arrays), but we now store FitResultData (scalars).
    # TODO: Update export_fit_results to work with FitResultData or reconstruct arrays on demand.
    if export_fits:
        # Combine measurement fits and level fits for this measurement/channel
        all_fits = {}
        measurement_fit = state.measurement_fits.get((measurement_id, channel))
        if measurement_fit:
            # Use -1 as level_id for measurement fits
            all_fits[(measurement_id, channel, -1)] = measurement_fit
        for (mid, ch, lvl_idx), fit in state.level_fits.items():
            if mid == measurement_id and ch == channel:
                all_fits[(mid, ch, lvl_idx)] = fit
        # Note: export_fit_results expects FitResult, but we have FitResultData now
        # This export functionality needs to be updated in a future task
        if all_fits:
            logger.warning("Fit export not yet updated for new FitResultData storage")

    return exported_files


def export_batch(
    state: SessionState,
    selections: list[tuple[int, int]],
    output_dir: Path,
    fmt: ExportFormat = ExportFormat.CSV,
    export_intensity: bool = True,
    export_levels: bool = True,
    export_groups: bool = True,
    export_fits: bool = True,
    bin_size_ms: float = 10.0,
    progress_callback: callable | None = None,
) -> list[Path]:
    """Export data for multiple measurements.

    Args:
        state: The session state containing all data.
        selections: List of (measurement_id, channel) tuples to export.
        output_dir: Directory to export files to.
        fmt: Export format.
        export_intensity: Whether to export intensity traces.
        export_levels: Whether to export levels.
        export_groups: Whether to export groups.
        export_fits: Whether to export fit results.
        bin_size_ms: Bin size for intensity traces.
        progress_callback: Optional callback(progress, message) for progress updates.

    Returns:
        List of paths to all exported files.
    """
    all_files: list[Path] = []
    total = len(selections)

    for idx, (measurement_id, channel) in enumerate(selections):
        if progress_callback:
            progress = (idx + 1) / total
            progress_callback(progress, f"Exporting measurement {measurement_id}...")

        files = export_all_measurement_data(
            state=state,
            measurement_id=measurement_id,
            channel=channel,
            output_dir=output_dir,
            fmt=fmt,
            export_intensity=export_intensity,
            export_levels=export_levels,
            export_groups=export_groups,
            export_fits=export_fits,
            bin_size_ms=bin_size_ms,
        )
        all_files.extend(files)

    if progress_callback:
        progress_callback(1.0, f"Exported {len(all_files)} files")

    logger.info(f"Batch export complete: {len(all_files)} files to {output_dir}")
    return all_files
